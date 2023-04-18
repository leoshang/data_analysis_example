import os
from openpyxl import load_workbook


def read_excel_file(relative_path):
    # Get the absolute path of the file
    abs_path = os.path.abspath(relative_path)

    # Load the workbook
    workbook = load_workbook(filename=abs_path, read_only=True)

    # Select the first worksheet
    worksheet = workbook.worksheets[0]

    # Read the data from the worksheet
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(row)

    return data


def main():
    # Define the relative path of the Excel file
    relative_path = "premier-league-2014-2015_analysis.xlsx"

    # Read the data from the Excel file
    data = read_excel_file(relative_path)

    # Print the data to the console
    for row in data:
        print(row)
