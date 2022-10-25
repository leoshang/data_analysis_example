# coding=utf-8
import xlsxwriter
import openpyxl
import json

from postgres_odds_analysis.config.configreader import ConfigurationReader
from postgres_odds_analysis.oddsitem import EuroOdds


class WilliamOddsAnalyser(object):
    def __init__(self):
        self.workbook = {}
        league = ConfigurationReader.get_league_name()
        season = ConfigurationReader.get_season()
        analysed_prefix = ConfigurationReader.get_analyse_prefix()
        output_file_name = analysed_prefix + league + season
        self.workbook = xlsxwriter.Workbook('%s.xlsx' % output_file_name)
        self.output_sheet = self.workbook.add_worksheet()
        self.output_sheet.set_default_row(25)
        self.row_count = 0
        #self.handicap_map = {u'两球': 2, u'球半/两球': 1.75, u'球半': 1.5, u'一球/球半': 1.25, u'一球': 1, u'半球/一球': 0.75, u'半球': 0.5,
        #                     u'平手/半球': 0.25, u'平手': 0}
        #self.handicap_map_rev = {2: u'两球', 1.75: u'球半/两球', 1.5: u'球半', 1.25: u'一球/球半', 1: u'一球', 0.75: u'半球/一球',
        #                         0.5: u'半球', 0.25: u'平手/半球', 0: u'平手'}

    def read_items(self):
        odds_data_file = ConfigurationReader.get_data_file()
        odds_reader = openpyxl.load_workbook(odds_data_file)

        # Get workbook active sheet object
        # from the active attribute
        odds_sheet = odds_reader.active
        return self.process_items(odds_sheet)

    def write_items(self, items):
        self.write_header()
        self.write(items)

    def write_header(self):
        self.output_sheet.write(self.row_count, 0, u'盘口')
        self.output_sheet.write(self.row_count, 1, u'赢平负')
        self.output_sheet.write(self.row_count, 2, u'赔率')
        self.output_sheet.write(self.row_count, 3, u'总场次')
        self.output_sheet.write(self.row_count, 4, u'赢场次')
        self.output_sheet.write(self.row_count, 5, u'平场次')
        self.output_sheet.write(self.row_count, 6, u'输场次')
        self.row_count += 1

    def process_items(self, odds_sheet):
        handicap_dict = dict()

        for i in range(2, odds_sheet.max_row + 1):
            handicap_draw_total = 0
            handicap_homewin_total = 0
            handicap_draw_homewin_total = 0

            handicap_draw_wins = 0
            handicap_draw_draws = 0
            handicap_draw_defeats = 0

            handicap_homewin_wins = 0
            handicap_homewin_draws = 0
            handicap_homewin_defeats = 0

            handicap_draw_homewin_wins = 0
            handicap_draw_homewin_draws = 0
            handicap_draw_homewin_defeats = 0

            handicap = odds_sheet.cell(i, 26).value
            if handicap:
                euro_odds = EuroOdds()
                euro_odds['open_home_win'] = odds_sheet.cell(i, 22).value
                euro_odds['open_draw'] = odds_sheet.cell(i, 23).value

                euro_odds['asian_start_home_wager'] = odds_sheet.cell(i, 25).value
                euro_odds['asian_start_handicap'] = odds_sheet.cell(i, 26).value

                euro_odds['host_goal'] = odds_sheet.cell(i, 43).value
                euro_odds['guest_goal'] = odds_sheet.cell(i, 44).value

                game_result = 0
                if int(euro_odds['host_goal']) > int(euro_odds['guest_goal']):
                    game_result = 1
                elif int(euro_odds['host_goal']) < int(euro_odds['guest_goal']):
                    game_result = -1

                handicap_draw = euro_odds['asian_start_handicap'] + '_' + u'平局赔率' + '_' + euro_odds['open_draw']
                print handicap_draw

                handicap_homewin = euro_odds['asian_start_handicap'] + '_' + u'主胜赔率' + '_' + euro_odds['open_home_win']
                print handicap_homewin

                handicap_draw_homewin = euro_odds['asian_start_handicap'] + '_' + euro_odds['open_draw'] + '_' + euro_odds['open_home_win']

                if handicap_draw in handicap_dict:
                    current_draw = handicap_dict[handicap_draw]
                    handicap_draw_total = int(current_draw['match_total']) + 1

                    if game_result == 0:
                        handicap_draw_draws = int(current_draw['match_draws']) + 1
                        handicap_draw_wins = int(current_draw['match_wins'])
                        handicap_draw_defeats = int(current_draw['match_defeats'])
                    elif game_result == 1:
                        handicap_draw_wins = int(current_draw['match_wins']) + 1
                        handicap_draw_draws = int(current_draw['match_draws'])
                        handicap_draw_defeats = int(current_draw['match_defeats'])
                    elif game_result == -1:
                        handicap_draw_defeats = int(current_draw['match_defeats']) + 1
                        handicap_draw_wins = int(current_draw['match_wins'])
                        handicap_draw_draws = int(current_draw['match_draws'])
                        
                    current_record = {'match_total': handicap_draw_total, 'match_wins': handicap_draw_wins,
                                      'match_draws': handicap_draw_draws, 'match_defeats': handicap_draw_defeats}
                else:
                    # current_draw = [euro_odds]
                    handicap_draw_total = 1
                    if game_result == 0:
                        handicap_draw_draws = 1
                    else:
                        handicap_draw_draws = 0
                    if game_result == 1:
                        handicap_draw_wins = 1
                    else:
                        handicap_draw_wins = 0
                    if game_result == -1:
                        handicap_draw_defeats = 1
                    else:
                        handicap_draw_defeats = 0
                    current_record = {'match_total': handicap_draw_total, 'match_wins': handicap_draw_wins, 'match_draws': handicap_draw_draws, 'match_defeats': handicap_draw_defeats}
                if current_record:
                    handicap_dict[handicap_draw] = current_record

                if handicap_homewin in handicap_dict:
                    # handicap_dict[handicap_draw].append()
                    current_homewin = handicap_dict[handicap_homewin]
                    handicap_homewin_total = int(current_homewin['match_total']) + 1
                    if game_result == 0:
                        handicap_homewin_draws = int(current_homewin['match_draws']) + 1
                        handicap_homewin_wins = int(current_homewin['match_wins'])
                        handicap_homewin_defeats = int(current_homewin['match_defeats'])
                    elif game_result == 1:
                        handicap_homewin_wins = int(current_homewin['match_wins']) + 1
                        handicap_homewin_draws = int(current_homewin['match_draws'])
                        handicap_homewin_defeats = int(current_homewin['match_defeats'])
                    elif game_result == -1:
                        handicap_homewin_defeats = int(current_homewin['match_defeats']) + 1
                        handicap_homewin_wins = int(current_homewin['match_wins'])
                        handicap_homewin_draws = int(current_homewin['match_draws'])
                    current_record_2 = {'match_total': handicap_homewin_total, 'match_wins': handicap_homewin_wins,
                                        'match_draws': handicap_homewin_draws, 'match_defeats': handicap_homewin_defeats}
                else:
                    # handicap_dict[handicap_draw] = [euro_odds]
                    handicap_homewin_total = 1
                    if game_result == 0:
                        handicap_homewin_draws = 1
                    else:
                        handicap_homewin_draws = 0

                    if game_result == 1:
                        handicap_homewin_wins = 1
                    else:
                        handicap_homewin_wins = 0

                    if game_result == -1:
                        handicap_homewin_defeats = 1
                    else:
                        handicap_homewin_defeats = 0

                    current_record_2 = {'match_total': handicap_homewin_total, 'match_wins': handicap_homewin_wins,
                                        'match_draws': handicap_homewin_draws, 'match_defeats': handicap_homewin_defeats}

                if current_record_2:
                    handicap_dict[handicap_homewin] = current_record_2

        return handicap_dict

    def write(self, handicap_dict):
        self.write_header()
        for key in sorted(handicap_dict):
            odds = handicap_dict[key]
            self.do_write(key, odds)
            # self.output_sheet.insert_rows(self.row_count)
            # self.row_count += 1

    def do_write(self, handicap, odd):
        handicap_arr = handicap.split('_')
        self.output_sheet.write(self.row_count, 0, handicap_arr[0])
        self.output_sheet.write(self.row_count, 1, handicap_arr[1])
        self.output_sheet.write(self.row_count, 2, handicap_arr[2])
        self.output_sheet.write(self.row_count, 3, odd['match_total'])
        self.output_sheet.write(self.row_count, 4, odd['match_wins'])
        self.output_sheet.write(self.row_count, 5, odd['match_draws'])
        self.output_sheet.write(self.row_count, 6, odd['match_defeats'])

        self.row_count += 1

    def close_writer(self):
        self.workbook.close()
